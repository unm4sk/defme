from rich.console import Console
from rich.progress import track
from rich.theme import Theme
from openpyxl import Workbook, load_workbook
import os
import time
from settings import spreadsheet, worksheet
from PyDictionary import PyDictionary

dictionary = PyDictionary()

custom_theme = Theme({'success': 'bold underline green', 'warning': 'underline yellow', 'error': 'bold red', 'error_api': 'blue'})
console = Console(theme=custom_theme)

os.system('cls' if os.name == 'nt' else 'clear')

#TODO implement undrelines if word appears in definition
#TODO add synonyms & antonyms
#TODO add examples <- create my own method 


workbook = load_workbook(f'{spreadsheet}.xlsx')
worksheet = workbook[worksheet]

possibles = ['Noun', 'Verb', 'Adjective', 'Adverb']
not_single = []
review = []
cur_pos = 2
rows_count = 1
ran = False

def county():
    global rows_count, ran
    cnt_track = 0
    for i in range(rows_count, worksheet.max_row+1):
        if cnt_track >=10: # checking for the end of spreadsheet
            rows_count -= 10
            break
        if (worksheet[f'A{i}'].value == None):
            cnt_track += 1
            rows_count += 1
        elif (worksheet[f'A{i}'].value != None):
            cnt_track = 0
            rows_count += 1
        time.sleep(0.1)
    console.print('Words counted!', style='success')
    ran = True
    return rows_count-1


def dict(i):
    result = ''
    definition = dictionary.meaning(f'{worksheet[f"A{i}"].value}', disable_errors=True)
    # print(definition)
    for i in possibles:
        try: 
            definition[i]
            if len(definition[i]) == 1:
                result += f'{i}: ' + definition[i][0] + '\n'
            elif len(definition[i]) >=3:
                result += f'{i}: '
                for pos in range(3):
                    result += f'{pos+1}) ' + definition[i][pos] + '\n'
            else:
                result += f'{i}: '
                for pos in range(len(definition[i])+1):
                    result += f'{pos+1}) ' + definition[i][pos] + '\n'
        except:
            pass
    return result.strip()

def search(beg=2,end=rows_count): # current position
    global cur_pos, review
    for i in track(range(beg, rows_count), description='Processing...'):
        try: 
            if len((worksheet[f'A{i}'].value).split()) > 1:
                not_single.append(f'A{i}')
        except:
            pass
        if (worksheet[f'A{i}'].value != None) and (worksheet[f'B{i}'].value == None):
            worksheet[f'B{i}'] = dict(i)
        elif (worksheet[f'A{i}'].value == None) and (worksheet[f'B{i}'].value != None): 
            worksheet[f'B{i}'].value = 'REVIEW!!!'
            review.append(f'A{i}:B{i}')
        cur_pos += 1
        time.sleep(0.1)


def main():
    global ran
    if not ran:
        with console.status('Counting all words in cells...', spinner='material'):
            county()
    search(beg=cur_pos)
    workbook.save(f'{spreadsheet}.xlsx')
    console.print('Program sucessfully executed!', style='success')
    if review:
        console.print('\nDuring the execution these cells missing word but having definition were found:\n', style='warning')
        print(', '.join(review))
    if not_single:
        console.print('\nDuring the execution these cells containing multiple words were found:\n', style='warning')
        print(', '.join(not_single))


if __name__ == '__main__':
    while True:
        try:
            main()
            break
        except:
            console.print("An error has occured!", style='error')
            break
